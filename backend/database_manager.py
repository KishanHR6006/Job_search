"""
Excel Database Manager
Tracks all applications with version control for tailored resumes
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging


class ApplicationDatabase:
    """
    Manages Excel spreadsheet for tracking job applications.
    Includes versioning for tailored resumes.
    """
    
    def __init__(self, db_path: str = "database/applications.xlsx"):
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Initialize or load existing database
        if not self.db_path.exists():
            self._create_new_database()
        
        self.df = pd.read_excel(self.db_path, sheet_name='Applications')
        
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
    
    def _create_new_database(self):
        """Create new Excel database with proper schema."""
        
        # Define columns
        columns = [
            'Application_ID',
            'Date_Applied',
            'Company',
            'Job_Title',
            'Job_URL',
            'Location',
            'Status',
            'Resume_Version',
            'Resume_Path',
            'Cover_Letter_Used',
            'Tailoring_Score',
            'Key_Matches',
            'Response_Date',
            'Response_Type',
            'Interview_Date',
            'Notes',
            'Follow_Up_Date'
        ]
        
        # Create empty DataFrame
        df = pd.DataFrame(columns=columns)
        
        # Create Excel writer with multiple sheets
        with pd.ExcelWriter(self.db_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Applications', index=False)
            
            # Add a summary sheet
            summary_df = pd.DataFrame({
                'Metric': [
                    'Total Applications',
                    'Applications This Week',
                    'Response Rate',
                    'Interview Rate',
                    'Average Tailoring Score'
                ],
                'Value': [0, 0, '0%', '0%', 0]
            })
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Add resume versions sheet
            versions_df = pd.DataFrame(columns=[
                'Version_ID',
                'Job_Hash',
                'Created_Date',
                'Base_Resume',
                'Company',
                'Title',
                'File_Path'
            ])
            versions_df.to_excel(writer, sheet_name='Resume_Versions', index=False)
        
        # Apply formatting
        self._apply_formatting()
        
        self.logger.info(f"Created new database: {self.db_path}")
    
    def _apply_formatting(self):
        """Apply professional Excel formatting."""
        
        wb = openpyxl.load_workbook(self.db_path)
        ws = wb['Applications']
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        column_widths = {
            'A': 15,  # Application_ID
            'B': 12,  # Date_Applied
            'C': 20,  # Company
            'D': 30,  # Job_Title
            'E': 50,  # Job_URL
            'F': 15,  # Location
            'G': 12,  # Status
            'H': 15,  # Resume_Version
            'I': 50,  # Resume_Path
            'J': 15,  # Cover_Letter_Used
            'K': 15,  # Tailoring_Score
            'L': 30,  # Key_Matches
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        wb.save(self.db_path)
    
    def add_application(self, application_data: Dict) -> str:
        """
        Add a new application to the database.
        
        Args:
            application_data: {
                'company': str,
                'job_title': str,
                'job_url': str,
                'location': str,
                'resume_version': str,
                'resume_path': str,
                'cover_letter_used': bool,
                'tailoring_score': int,
                'key_matches': List[str],
                'notes': str (optional)
            }
        
        Returns:
            Application ID
        """
        
        # Generate application ID
        app_id = f"APP_{datetime.now().strftime('%Y%m%d')}_{len(self.df) + 1:04d}"
        
        # Create new row
        new_row = {
            'Application_ID': app_id,
            'Date_Applied': datetime.now().strftime('%Y-%m-%d'),
            'Company': application_data['company'],
            'Job_Title': application_data['job_title'],
            'Job_URL': application_data['job_url'],
            'Location': application_data.get('location', ''),
            'Status': 'Applied',
            'Resume_Version': application_data['resume_version'],
            'Resume_Path': application_data['resume_path'],
            'Cover_Letter_Used': 'Yes' if application_data.get('cover_letter_used') else 'No',
            'Tailoring_Score': application_data.get('tailoring_score', 0),
            'Key_Matches': ', '.join(application_data.get('key_matches', [])),
            'Response_Date': '',
            'Response_Type': '',
            'Interview_Date': '',
            'Notes': application_data.get('notes', ''),
            'Follow_Up_Date': ''
        }
        
        # Append to DataFrame
        self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)
        
        # Save to Excel
        self._save()
        
        self.logger.info(f"Added application: {app_id} - {application_data['company']}")
        
        return app_id
    
    def update_status(self, app_id: str, status: str, notes: str = ""):
        """Update application status."""
        
        mask = self.df['Application_ID'] == app_id
        self.df.loc[mask, 'Status'] = status
        
        if notes:
            existing_notes = self.df.loc[mask, 'Notes'].values[0]
            updated_notes = f"{existing_notes}\n{datetime.now().strftime('%Y-%m-%d')}: {notes}"
            self.df.loc[mask, 'Notes'] = updated_notes
        
        if status == 'Response Received':
            self.df.loc[mask, 'Response_Date'] = datetime.now().strftime('%Y-%m-%d')
        
        self._save()
        self.logger.info(f"Updated {app_id}: {status}")
    
    def add_resume_version(self, version_data: Dict) -> str:
        """
        Track a new tailored resume version.
        
        Args:
            version_data: {
                'job_hash': str,
                'base_resume': str,
                'company': str,
                'title': str,
                'file_path': str
            }
        
        Returns:
            Version ID
        """
        
        version_id = f"V_{version_data['job_hash']}"
        
        # Load resume versions sheet
        versions_df = pd.read_excel(self.db_path, sheet_name='Resume_Versions')
        
        new_version = {
            'Version_ID': version_id,
            'Job_Hash': version_data['job_hash'],
            'Created_Date': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'Base_Resume': version_data['base_resume'],
            'Company': version_data['company'],
            'Title': version_data['title'],
            'File_Path': version_data['file_path']
        }
        
        versions_df = pd.concat([versions_df, pd.DataFrame([new_version])], ignore_index=True)
        
        # Save
        with pd.ExcelWriter(self.db_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            versions_df.to_excel(writer, sheet_name='Resume_Versions', index=False)
        
        self.logger.info(f"Added resume version: {version_id}")
        
        return version_id
    
    def get_applications_by_status(self, status: str) -> pd.DataFrame:
        """Get all applications with specific status."""
        return self.df[self.df['Status'] == status]
    
    def get_summary_stats(self) -> Dict:
        """Generate summary statistics."""
        
        total = len(self.df)
        
        # This week
        today = datetime.now()
        week_start = today - pd.Timedelta(days=7)
        this_week = len(self.df[pd.to_datetime(self.df['Date_Applied']) >= week_start])
        
        # Response rate
        responses = len(self.df[self.df['Response_Date'].notna()])
        response_rate = (responses / total * 100) if total > 0 else 0
        
        # Interview rate
        interviews = len(self.df[self.df['Interview_Date'].notna()])
        interview_rate = (interviews / total * 100) if total > 0 else 0
        
        # Average tailoring score
        avg_score = self.df['Tailoring_Score'].mean() if total > 0 else 0
        
        return {
            'total_applications': total,
            'this_week': this_week,
            'response_rate': f"{response_rate:.1f}%",
            'interview_rate': f"{interview_rate:.1f}%",
            'avg_tailoring_score': round(avg_score, 1)
        }
    
    def update_summary_sheet(self):
        """Update the summary sheet with latest stats."""
        
        stats = self.get_summary_stats()
        
        summary_df = pd.DataFrame({
            'Metric': [
                'Total Applications',
                'Applications This Week',
                'Response Rate',
                'Interview Rate',
                'Average Tailoring Score'
            ],
            'Value': [
                stats['total_applications'],
                stats['this_week'],
                stats['response_rate'],
                stats['interview_rate'],
                stats['avg_tailoring_score']
            ]
        })
        
        with pd.ExcelWriter(self.db_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _save(self):
        """Save DataFrame to Excel."""
        with pd.ExcelWriter(self.db_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            self.df.to_excel(writer, sheet_name='Applications', index=False)
        
        self.update_summary_sheet()
        self._apply_formatting()


# Example usage
if __name__ == "__main__":
    db = ApplicationDatabase()
    
    # Add sample application
    app_data = {
        'company': 'Tech Startup Inc',
        'job_title': 'Senior Backend Engineer',
        'job_url': 'https://linkedin.com/jobs/123456',
        'location': 'Remote',
        'resume_version': 'V_abc12345',
        'resume_path': 'resumes/tailored/resume_TechStartup_abc12345.pdf',
        'cover_letter_used': True,
        'tailoring_score': 87,
        'key_matches': ['Python', 'FastAPI', 'Microservices', 'AWS'],
        'notes': 'Found via LinkedIn recommendation'
    }
    
    app_id = db.add_application(app_data)
    print(f"Created application: {app_id}")
    
    # Print summary
    print("\nSummary Stats:")
    print(db.get_summary_stats())
